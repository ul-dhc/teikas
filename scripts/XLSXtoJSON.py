import argparse, hashlib, json
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
import openpyxl

REQUIRED = {
    "Legends": ["legend_id", "volume", "chapter_lv", "chapter_de",
        "title_lv", "title_de", "collector_id", "narrator_id", "place_id",
        "original_metadata", "comments", "latitude_original",
        "longitude_original", "google_maps_original", "text_lv", "text_de",
        "notes", "primary_source_raw", "secondary_source_raw"],
    "Collectors": ["collector_id", "full_name", "gender", "legend_count", "notes"],
    "Narrators": ["narrator_id", "full_name", "gender", "legend_count", "notes"],
    "Places": ["place_id", "name", "latitude", "longitude",
        "coordinate_status", "coordinate_variants", "legend_count", "notes"],
    "Sources": ["source_id", "code", "title", "authors", "category",
        "category_original", "year", "publisher", "publication_place", "url",
        "original_label", "include_in_legend_site", "relation_count", "notes"],
    "LegendSources": ["relation_id", "legend_id", "source_id", "role",
        "raw_citation", "mapping_status", "notes"],
}

def rows(ws):
    values = ws.iter_rows(values_only=True)
    headers = [str(x).strip() for x in next(values)]
    if headers != REQUIRED[ws.title]:
        raise ValueError(f"Unexpected headers in {ws.title}: {headers}")
    for values_row in values:
        if any(v is not None and str(v).strip() for v in values_row):
            yield {k: (None if v is None or (isinstance(v, str) and not v.strip())
                    else v) for k, v in zip(headers, values_row)}

def coordinates(lat, lon):
    if (lat is None) != (lon is None):
        raise ValueError("Incomplete coordinate pair")
    if lat is None:
        return None
    lat, lon = float(lat), float(lon)
    if not (-90 <= lat <= 90 and -180 <= lon <= 180):
        raise ValueError(f"Invalid coordinates: {lat}, {lon}")
    return {"latitude": lat, "longitude": lon}

def unique(records, key, label):
    ids = [r[key] for r in records]
    if any(x is None for x in ids) or len(ids) != len(set(ids)):
        raise ValueError(f"Missing or duplicate {label} IDs")
    return set(ids)

def dump(path, value):
    # The serialized text already uses explicit LF separators; avoiding the
    # Python 3.10+ ``newline`` argument keeps the converter compatible with 3.9.
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2) + "\n",
                    encoding="utf-8")

def main(source, output):
    wb = openpyxl.load_workbook(source, read_only=True, data_only=True)
    for sheet in REQUIRED:
        if sheet not in wb.sheetnames:
            raise ValueError(f"Missing sheet: {sheet}")

    raw = {name: list(rows(wb[name])) for name in REQUIRED}
    collector_ids = unique(raw["Collectors"], "collector_id", "collector")
    narrator_ids = unique(raw["Narrators"], "narrator_id", "narrator")
    place_ids = unique(raw["Places"], "place_id", "place")
    source_ids = unique(raw["Sources"], "source_id", "source")
    legend_ids = unique(raw["Legends"], "legend_id", "legend")
    unique(raw["LegendSources"], "relation_id", "relationship")

    legends = []
    for r in raw["Legends"]:
        if r["collector_id"] not in collector_ids:
            raise ValueError(f"Unknown collector: {r['collector_id']}")
        if r["narrator_id"] is not None and r["narrator_id"] not in narrator_ids:
            raise ValueError(f"Unknown narrator: {r['narrator_id']}")
        if r["place_id"] is not None and r["place_id"] not in place_ids:
            raise ValueError(f"Unknown place: {r['place_id']}")
        legends.append({
            "id": str(r["legend_id"]), "volume": str(r["volume"]),
            "chapter": {"lv": r["chapter_lv"], "de": r["chapter_de"]},
            "title": {"lv": r["title_lv"], "de": r["title_de"]},
            "text": {"lv": r["text_lv"], "de": r["text_de"]},
            "collectorId": r["collector_id"], "narratorId": r["narrator_id"],
            "placeId": r["place_id"], "originalMetadata": r["original_metadata"],
            "comments": r["comments"], "notes": r["notes"],
            "originalCoordinates": coordinates(r["latitude_original"], r["longitude_original"]),
            "primarySourceRaw": r["primary_source_raw"],
            "secondarySourceRaw": r["secondary_source_raw"],
        })

    collector_counts = Counter(x["collectorId"] for x in legends)
    narrator_counts = Counter(x["narratorId"] for x in legends if x["narratorId"])
    place_counts = Counter(x["placeId"] for x in legends if x["placeId"])

    people = lambda sheet, prefix, counts: [{
        "id": r[f"{prefix}_id"], "fullName": r["full_name"],
        "gender": r["gender"], "notes": r["notes"],
        "legendCount": counts[r[f"{prefix}_id"]]
    } for r in raw[sheet]]

    places = [{
        "id": r["place_id"], "name": r["name"],
        "coordinates": coordinates(r["latitude"], r["longitude"]),
        "coordinateStatus": r["coordinate_status"],
        "coordinateVariantsRaw": r["coordinate_variants"], "notes": r["notes"],
        "legendCount": place_counts[r["place_id"]]
    } for r in raw["Places"]]

    links = []
    seen_tuples = set()
    for r in raw["LegendSources"]:
        if r["legend_id"] not in legend_ids or r["source_id"] not in source_ids:
            raise ValueError(f"Broken source relationship: {r['relation_id']}")
        signature = (r["legend_id"], r["source_id"], r["role"], r["raw_citation"])
        if signature in seen_tuples:
            raise ValueError(f"Duplicate source relationship: {signature}")
        seen_tuples.add(signature)
        links.append({"id": r["relation_id"], "legendId": r["legend_id"],
            "sourceId": r["source_id"], "role": r["role"],
            "rawCitation": r["raw_citation"], "mappingStatus": r["mapping_status"],
            "notes": r["notes"]})

    source_counts = Counter(x["sourceId"] for x in links)
    sources = [{
        "id": r["source_id"], "code": r["code"], "title": r["title"],
        "authorsRaw": r["authors"], "category": r["category"],
        "categoryOriginal": r["category_original"], "year": r["year"],
        "publisher": r["publisher"], "publicationPlace": r["publication_place"],
        "url": r["url"], "originalLabel": r["original_label"],
        "notes": r["notes"], "relationCount": source_counts[r["source_id"]]
    } for r in raw["Sources"]]

    datasets = {
        "legends.json": legends,
        "collectors.json": people("Collectors", "collector", collector_counts),
        "narrators.json": people("Narrators", "narrator", narrator_counts),
        "places.json": places, "sources.json": sources,
        "legend-sources.json": links,
    }

    output.mkdir(parents=True, exist_ok=True)
    for filename, records in datasets.items():
        dump(output / filename, sorted(records, key=lambda x: x["id"]))
    manifest = {"schemaVersion": 1,
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "sourceFile": source.name,
        "sourceSha256": hashlib.sha256(source.read_bytes()).hexdigest(),
        "recordCounts": {name: len(value) for name, value in datasets.items()}}
    dump(output / "dataset-manifest.json", manifest)

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("source", type=Path)
    parser.add_argument("output", type=Path)
    args = parser.parse_args()
    main(args.source, args.output)
