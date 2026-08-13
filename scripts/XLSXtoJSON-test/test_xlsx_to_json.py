"""Regression and validation tests for XLSXtoJSON.py."""

from __future__ import annotations

import hashlib
import importlib.util
import json
import os
import tempfile
import unittest
from collections import Counter
from copy import deepcopy
from pathlib import Path

import openpyxl


TEST_DIR = Path(__file__).resolve().parent
SCRIPTS_DIR = TEST_DIR.parent
PROJECT_DIR = SCRIPTS_DIR.parent
CONVERTER_PATH = SCRIPTS_DIR / "XLSXtoJSON.py"
SOURCE_XLSX = Path(os.environ.get("XLSX_SOURCE", PROJECT_DIR / "latvian_legends_master_data.xlsx"))
JSON_OUTPUT = Path(os.environ.get("JSON_OUTPUT", PROJECT_DIR / "teikas_json"))

spec = importlib.util.spec_from_file_location("xlsx_to_json", CONVERTER_PATH)
converter = importlib.util.module_from_spec(spec)
assert spec.loader is not None
spec.loader.exec_module(converter)


def load_json(filename: str):
    path = JSON_OUTPUT / filename
    if not path.is_file():
        raise AssertionError(
            f"Missing generated file '{path}'. Run XLSXtoJSON.py before the tests."
        )
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise AssertionError(f"'{path}' is not valid JSON: {exc}") from exc


def read_sheet(workbook, sheet_name: str) -> list[dict]:
    sheet = workbook[sheet_name]
    values = sheet.iter_rows(values_only=True)
    headers = [str(value).strip() for value in next(values)]
    records = []
    for row_number, row in enumerate(values, start=2):
        if not any(value is not None and str(value).strip() for value in row):
            continue
        record = {
            key: None if value is None or (isinstance(value, str) and not value.strip()) else value
            for key, value in zip(headers, row)
        }
        record["__row__"] = row_number
        records.append(record)
    return records


def coordinate_object(latitude, longitude):
    if latitude is None and longitude is None:
        return None
    return {"latitude": float(latitude), "longitude": float(longitude)}


def by_id(records: list[dict], label: str) -> dict:
    if not isinstance(records, list):
        raise AssertionError(f"{label} JSON must contain a top-level array.")
    result = {}
    for index, record in enumerate(records):
        if not isinstance(record, dict):
            raise AssertionError(f"{label}[{index}] must be a JSON object.")
        object_id = record.get("id")
        if not object_id:
            raise AssertionError(f"{label}[{index}] has no non-empty 'id'.")
        if object_id in result:
            raise AssertionError(
                f"Duplicate ID '{object_id}' found in {label} JSON. IDs must be unique."
            )
        result[object_id] = record
    return result


def assert_record_equal(testcase, label, object_id, expected, actual):
    testcase.assertEqual(
        set(expected), set(actual),
        f"{label} '{object_id}' has different fields. "
        f"Expected {sorted(expected)}; found {sorted(actual)}.",
    )
    for field, expected_value in expected.items():
        testcase.assertEqual(
            expected_value, actual[field],
            f"{label} '{object_id}' field '{field}' differs: "
            f"XLSX implies {expected_value!r}, JSON contains {actual[field]!r}.",
        )


class PublishedContentTests(unittest.TestCase):
    """Compare the complete checked-in JSON publication with the XLSX source."""

    @classmethod
    def setUpClass(cls):
        if not SOURCE_XLSX.is_file():
            raise AssertionError(f"Source workbook not found: '{SOURCE_XLSX}'.")
        cls.workbook = openpyxl.load_workbook(SOURCE_XLSX, read_only=True, data_only=True)
        cls.raw = {name: read_sheet(cls.workbook, name) for name in converter.REQUIRED}
        cls.data = {
            "legends": load_json("legends.json"),
            "collectors": load_json("collectors.json"),
            "narrators": load_json("narrators.json"),
            "places": load_json("places.json"),
            "sources": load_json("sources.json"),
            "links": load_json("legend-sources.json"),
        }
        cls.index = {name: by_id(records, name) for name, records in cls.data.items()}

    @classmethod
    def tearDownClass(cls):
        cls.workbook.close()

    def test_required_sheet_headers_are_unchanged(self):
        for sheet_name, required_headers in converter.REQUIRED.items():
            with self.subTest(sheet=sheet_name):
                actual = [str(cell.value).strip() for cell in self.workbook[sheet_name][1]]
                self.assertEqual(
                    required_headers, actual,
                    f"Sheet '{sheet_name}' headers changed. Expected {required_headers}; found {actual}.",
                )

    def test_every_legend_matches_its_xlsx_row(self):
        expected = {}
        for row in self.raw["Legends"]:
            object_id = str(row["legend_id"])
            expected[object_id] = {
                "id": object_id,
                "volume": str(row["volume"]),
                "chapter": {"lv": row["chapter_lv"], "de": row["chapter_de"]},
                "title": {"lv": row["title_lv"], "de": row["title_de"]},
                "text": {"lv": row["text_lv"], "de": row["text_de"]},
                "collectorId": row["collector_id"],
                "narratorId": row["narrator_id"],
                "placeId": row["place_id"],
                "originalMetadata": row["original_metadata"],
                "comments": row["comments"],
                "notes": row["notes"],
                "originalCoordinates": coordinate_object(
                    row["latitude_original"], row["longitude_original"]
                ),
                "primarySourceRaw": row["primary_source_raw"],
                "secondarySourceRaw": row["secondary_source_raw"],
            }
        self.assertEqual(
            set(expected), set(self.index["legends"]),
            "Legend IDs differ between the XLSX sheet and legends.json.",
        )
        for object_id, record in expected.items():
            assert_record_equal(self, "Legend", object_id, record, self.index["legends"][object_id])

    def test_every_collector_and_narrator_matches_xlsx(self):
        legends = self.data["legends"]
        cases = (
            ("Collectors", "collectors", "collector", "collectorId"),
            ("Narrators", "narrators", "narrator", "narratorId"),
        )
        for sheet, dataset, prefix, link_field in cases:
            counts = Counter(x[link_field] for x in legends if x[link_field] is not None)
            expected = {}
            for row in self.raw[sheet]:
                object_id = row[f"{prefix}_id"]
                expected[object_id] = {
                    "id": object_id,
                    "fullName": row["full_name"],
                    "gender": row["gender"],
                    "notes": row["notes"],
                    "legendCount": counts[object_id],
                }
            with self.subTest(dataset=dataset):
                self.assertEqual(
                    set(expected), set(self.index[dataset]),
                    f"IDs differ between the {sheet} sheet and {dataset}.json.",
                )
                for object_id, record in expected.items():
                    assert_record_equal(self, prefix.title(), object_id, record, self.index[dataset][object_id])

    def test_every_place_matches_its_xlsx_row(self):
        counts = Counter(x["placeId"] for x in self.data["legends"] if x["placeId"] is not None)
        expected = {}
        for row in self.raw["Places"]:
            object_id = row["place_id"]
            expected[object_id] = {
                "id": object_id,
                "name": row["name"],
                "coordinates": coordinate_object(row["latitude"], row["longitude"]),
                "coordinateStatus": row["coordinate_status"],
                "coordinateVariantsRaw": row["coordinate_variants"],
                "notes": row["notes"],
                "legendCount": counts[object_id],
            }
        self.assertEqual(set(expected), set(self.index["places"]), "Place IDs differ between XLSX and JSON.")
        for object_id, record in expected.items():
            assert_record_equal(self, "Place", object_id, record, self.index["places"][object_id])

    def test_every_source_and_relationship_matches_xlsx(self):
        expected_links = {}
        for row in self.raw["LegendSources"]:
            object_id = row["relation_id"]
            expected_links[object_id] = {
                "id": object_id,
                "legendId": row["legend_id"],
                "sourceId": row["source_id"],
                "role": row["role"],
                "rawCitation": row["raw_citation"],
                "mappingStatus": row["mapping_status"],
                "notes": row["notes"],
            }
        self.assertEqual(
            set(expected_links), set(self.index["links"]),
            "Relationship IDs differ between LegendSources and legend-sources.json.",
        )
        for object_id, record in expected_links.items():
            assert_record_equal(self, "Legend-source relationship", object_id, record, self.index["links"][object_id])

        counts = Counter(x["sourceId"] for x in self.data["links"])
        expected_sources = {}
        for row in self.raw["Sources"]:
            object_id = row["source_id"]
            expected_sources[object_id] = {
                "id": object_id,
                "code": row["code"],
                "title": row["title"],
                "authorsRaw": row["authors"],
                "category": row["category"],
                "categoryOriginal": row["category_original"],
                "year": row["year"],
                "publisher": row["publisher"],
                "publicationPlace": row["publication_place"],
                "url": row["url"],
                "originalLabel": row["original_label"],
                "notes": row["notes"],
                "relationCount": counts[object_id],
            }
        self.assertEqual(set(expected_sources), set(self.index["sources"]), "Source IDs differ between XLSX and JSON.")
        for object_id, record in expected_sources.items():
            assert_record_equal(self, "Source", object_id, record, self.index["sources"][object_id])

    def test_all_foreign_keys_resolve(self):
        for legend in self.data["legends"]:
            self.assertIn(
                legend["collectorId"], self.index["collectors"],
                f"Legend '{legend['id']}' refers to unknown collector '{legend['collectorId']}'.",
            )
            if legend["narratorId"] is not None:
                self.assertIn(
                    legend["narratorId"], self.index["narrators"],
                    f"Legend '{legend['id']}' refers to unknown narrator '{legend['narratorId']}'.",
                )
            if legend["placeId"] is not None:
                self.assertIn(
                    legend["placeId"], self.index["places"],
                    f"Legend '{legend['id']}' refers to unknown place '{legend['placeId']}'.",
                )
        for link in self.data["links"]:
            self.assertIn(
                link["legendId"], self.index["legends"],
                f"Relationship '{link['id']}' refers to unknown legend '{link['legendId']}'.",
            )
            self.assertIn(
                link["sourceId"], self.index["sources"],
                f"Relationship '{link['id']}' refers to unknown source '{link['sourceId']}'.",
            )

    def test_relationship_tuples_are_unique(self):
        seen = set()
        for link in self.data["links"]:
            signature = (link["legendId"], link["sourceId"], link["role"], link["rawCitation"])
            self.assertNotIn(
                signature, seen,
                f"Duplicate legend-source relationship found: {signature!r}.",
            )
            seen.add(signature)

    def test_all_published_counts_equal_actual_links(self):
        cases = (
            ("collectors", "collectorId", "legendCount"),
            ("narrators", "narratorId", "legendCount"),
            ("places", "placeId", "legendCount"),
        )
        for dataset, link_field, count_field in cases:
            actual = Counter(x[link_field] for x in self.data["legends"] if x[link_field] is not None)
            for entity in self.data[dataset]:
                with self.subTest(dataset=dataset, object_id=entity["id"]):
                    self.assertEqual(
                        actual[entity["id"]], entity[count_field],
                        f"{dataset[:-1].title()} '{entity['id']}' says {count_field}="
                        f"{entity[count_field]}, but exactly {actual[entity['id']]} legends link to it.",
                    )
        source_counts = Counter(x["sourceId"] for x in self.data["links"])
        for source in self.data["sources"]:
            self.assertEqual(
                source_counts[source["id"]], source["relationCount"],
                f"Source '{source['id']}' says relationCount={source['relationCount']}, "
                f"but exactly {source_counts[source['id']]} relationships link to it.",
            )

    def test_records_are_sorted_by_id(self):
        for dataset, records in self.data.items():
            ids = [record["id"] for record in records]
            self.assertEqual(ids, sorted(ids), f"{dataset} records are not sorted by ID.")

    def test_manifest_matches_source_and_json(self):
        manifest = load_json("dataset-manifest.json")
        self.assertEqual(1, manifest.get("schemaVersion"), "Manifest schemaVersion must be 1.")
        self.assertEqual(SOURCE_XLSX.name, manifest.get("sourceFile"), "Manifest sourceFile is incorrect.")
        expected_hash = hashlib.sha256(SOURCE_XLSX.read_bytes()).hexdigest()
        self.assertEqual(
            expected_hash, manifest.get("sourceSha256"),
            "Manifest checksum does not match the current XLSX file; regenerate the JSON.",
        )
        expected_counts = {
            "legends.json": len(self.data["legends"]),
            "collectors.json": len(self.data["collectors"]),
            "narrators.json": len(self.data["narrators"]),
            "places.json": len(self.data["places"]),
            "sources.json": len(self.data["sources"]),
            "legend-sources.json": len(self.data["links"]),
        }
        self.assertEqual(
            expected_counts, manifest.get("recordCounts"),
            "Manifest recordCounts do not match the generated JSON files.",
        )


def minimal_records() -> dict[str, list[dict]]:
    """Return one valid, linked record for every required worksheet."""
    return {
        "Legends": [{
            "legend_id": "130101001", "volume": "13", "chapter_lv": "Nodaļa",
            "chapter_de": "Kapitel", "title_lv": "Virsraksts", "title_de": "Titel",
            "collector_id": "collector-0001", "narrator_id": "narrator-0001",
            "place_id": "place-0001", "original_metadata": "Metadati",
            "comments": None, "latitude_original": 56.95, "longitude_original": 24.10,
            "google_maps_original": "https://example.test/map", "text_lv": "Teksts",
            "text_de": "Text", "notes": None, "primary_source_raw": "SRC 1",
            "secondary_source_raw": None,
        }],
        "Collectors": [{"collector_id": "collector-0001", "full_name": "Kolektors",
            "gender": "M", "legend_count": 1, "notes": None}],
        "Narrators": [{"narrator_id": "narrator-0001", "full_name": "Teicēja",
            "gender": "F", "legend_count": 1, "notes": None}],
        "Places": [{"place_id": "place-0001", "name": "Rīga", "latitude": 56.95,
            "longitude": 24.10, "coordinate_status": "consistent",
            "coordinate_variants": "56.950000, 24.100000 (1)", "legend_count": 1,
            "notes": None}],
        "Sources": [{"source_id": "source-0001", "code": "SRC", "title": "Avots",
            "authors": None, "category": "published", "category_original": "publicēts",
            "year": "1926", "publisher": None, "publication_place": None, "url": None,
            "original_label": "SRC", "include_in_legend_site": True,
            "relation_count": 1, "notes": None}],
        "LegendSources": [{"relation_id": "relation-00001", "legend_id": "130101001",
            "source_id": "source-0001", "role": "primary", "raw_citation": "SRC 1",
            "mapping_status": "mapped", "notes": None}],
    }


def write_fixture(path: Path, records: dict[str, list[dict]], omit_sheet=None, header_change=None):
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    for sheet_name, headers in converter.REQUIRED.items():
        if sheet_name == omit_sheet:
            continue
        sheet = workbook.create_sheet(sheet_name)
        actual_headers = list(headers)
        if header_change and header_change[0] == sheet_name:
            actual_headers[header_change[1]] = header_change[2]
        sheet.append(actual_headers)
        for record in records[sheet_name]:
            sheet.append([record.get(header) for header in headers])
    workbook.save(path)
    workbook.close()


class ConverterValidationTests(unittest.TestCase):
    """Use small real XLSX fixtures to verify concise converter failures."""

    def run_fixture(self, records=None, **write_options):
        temporary = tempfile.TemporaryDirectory()
        self.addCleanup(temporary.cleanup)
        root = Path(temporary.name)
        source = root / "fixture.xlsx"
        output = root / "json"
        write_fixture(source, records or minimal_records(), **write_options)
        converter.main(source, output)
        return output

    def test_valid_fixture_converts_and_preserves_unicode_and_nulls(self):
        output = self.run_fixture()
        legend = json.loads((output / "legends.json").read_text(encoding="utf-8"))[0]
        self.assertEqual("Virsraksts", legend["title"]["lv"], "Latvian Unicode text was altered.")
        self.assertIsNone(legend["comments"], "A blank XLSX cell must become JSON null.")
        self.assertEqual(
            {"latitude": 56.95, "longitude": 24.1}, legend["originalCoordinates"],
            "Coordinate values were not preserved.",
        )

    def test_duplicate_ids_are_rejected_for_every_entity_type(self):
        cases = (
            ("Collectors", "collector_id", "collector"),
            ("Narrators", "narrator_id", "narrator"),
            ("Places", "place_id", "place"),
            ("Sources", "source_id", "source"),
            ("Legends", "legend_id", "legend"),
            ("LegendSources", "relation_id", "relationship"),
        )
        for sheet, id_field, label in cases:
            with self.subTest(sheet=sheet):
                records = minimal_records()
                duplicate = deepcopy(records[sheet][0])
                records[sheet].append(duplicate)
                with self.assertRaisesRegex(ValueError, rf"Missing or duplicate {label} IDs"):
                    self.run_fixture(records)

    def test_missing_id_is_rejected(self):
        records = minimal_records()
        records["Collectors"][0]["collector_id"] = None
        with self.assertRaisesRegex(ValueError, "Missing or duplicate collector IDs"):
            self.run_fixture(records)

    def test_broken_legend_foreign_keys_are_rejected(self):
        cases = (
            ("collector_id", "collector-missing", "Unknown collector: collector-missing"),
            ("narrator_id", "narrator-missing", "Unknown narrator: narrator-missing"),
            ("place_id", "place-missing", "Unknown place: place-missing"),
        )
        for field, bad_id, message in cases:
            with self.subTest(field=field):
                records = minimal_records()
                records["Legends"][0][field] = bad_id
                with self.assertRaisesRegex(ValueError, message):
                    self.run_fixture(records)

    def test_broken_source_relationship_ends_are_rejected(self):
        for field, bad_id in (("legend_id", "999999999"), ("source_id", "source-missing")):
            with self.subTest(field=field):
                records = minimal_records()
                records["LegendSources"][0][field] = bad_id
                with self.assertRaisesRegex(ValueError, "Broken source relationship: relation-00001"):
                    self.run_fixture(records)

    def test_duplicate_relationship_tuple_is_rejected(self):
        records = minimal_records()
        duplicate = deepcopy(records["LegendSources"][0])
        duplicate["relation_id"] = "relation-00002"
        records["LegendSources"].append(duplicate)
        with self.assertRaisesRegex(ValueError, "Duplicate source relationship"):
            self.run_fixture(records)

    def test_incomplete_and_out_of_range_coordinates_are_rejected(self):
        cases = (
            ("Legends", "longitude_original", None, "Incomplete coordinate pair"),
            ("Places", "longitude", None, "Incomplete coordinate pair"),
            ("Legends", "latitude_original", 91, "Invalid coordinates"),
            ("Places", "longitude", 181, "Invalid coordinates"),
        )
        for sheet, field, value, message in cases:
            with self.subTest(sheet=sheet, field=field):
                records = minimal_records()
                records[sheet][0][field] = value
                with self.assertRaisesRegex(ValueError, message):
                    self.run_fixture(records)

    def test_missing_required_sheet_is_rejected(self):
        with self.assertRaisesRegex(ValueError, "Missing sheet: Sources"):
            self.run_fixture(omit_sheet="Sources")

    def test_changed_header_is_rejected_with_sheet_name(self):
        with self.assertRaisesRegex(ValueError, "Unexpected headers in Legends"):
            self.run_fixture(header_change=("Legends", 0, "changed_legend_id"))

    def test_fixture_counts_are_derived_not_copied_from_excel_formulas(self):
        records = minimal_records()
        second_legend = deepcopy(records["Legends"][0])
        second_legend["legend_id"] = "130101002"
        records["Legends"].append(second_legend)
        records["Collectors"][0]["legend_count"] = 999
        records["Narrators"][0]["legend_count"] = 999
        records["Places"][0]["legend_count"] = 999
        records["Sources"][0]["relation_count"] = 999
        output = self.run_fixture(records)
        collectors = json.loads((output / "collectors.json").read_text(encoding="utf-8"))
        narrators = json.loads((output / "narrators.json").read_text(encoding="utf-8"))
        places = json.loads((output / "places.json").read_text(encoding="utf-8"))
        sources = json.loads((output / "sources.json").read_text(encoding="utf-8"))
        self.assertEqual(2, collectors[0]["legendCount"], "Collector count must be derived from legends.")
        self.assertEqual(2, narrators[0]["legendCount"], "Narrator count must be derived from legends.")
        self.assertEqual(2, places[0]["legendCount"], "Place count must be derived from legends.")
        self.assertEqual(1, sources[0]["relationCount"], "Source count must be derived from relationships.")


if __name__ == "__main__":
    unittest.main(verbosity=2)
